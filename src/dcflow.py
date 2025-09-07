from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import cvxpy as cp
import time
from pathlib import Path

input_file = Path("inputs/DCPOWERFLOW.xlsx")
output_file = Path("outputs/resultsdc.xlsx")

start_time = time.time()
wb = load_workbook(input_file)

sheet1 = wb['Sheet1']
Pg_max = [cell.value for row in sheet1['C2:C5'] for cell in row]
Cg = [cell.value for row in sheet1['D2:D5'] for cell in row]
Pd_max = [cell.value for row in sheet1['H2:H5'] for cell in row]
Bd = [cell.value for row in sheet1['I2:I5'] for cell in row]

sheet2 = wb['Sheet2']
num_buses = sheet2['B1'].value

connection_nodes = {i+1: [sheet2.cell(row=5+i, column=2).value, sheet2.cell(row=5+i, column=3).value]
                    for i in range(num_buses)}

generator_nodes = {i+1: [] for i in range(num_buses)}
demand_nodes = {i+1: [] for i in range(num_buses)}

susceptance = {i+1: [0]*4 for i in range(num_buses)}
for i in range(3, 3 + num_buses):
    for j in range(7, 7 + num_buses):
        row_bus = sheet2.cell(row=i, column=6).value
        col_bus = sheet2.cell(row=2, column=j).value
        if row_bus != col_bus:
            susceptance[row_bus][col_bus] = sheet2.cell(row=i, column=j).value

pmax = {i+1: [0]*4 for i in range(num_buses)}
for i in range(10, 10 + num_buses):
    for j in range(7, 7 + num_buses):
        row_bus = sheet2.cell(row=i, column=6).value
        col_bus = sheet2.cell(row=9, column=j).value
        if row_bus != col_bus:
            pmax[row_bus][col_bus] = sheet2.cell(row=i, column=j).value

gen_index = 0
for i in range(2, 2 + num_buses + 1):
    val = sheet1.cell(row=i, column=5).value
    if val in [1,2,3]:
        generator_nodes[val].append(gen_index)
        gen_index += 1

dem_index = 0
for i in range(2, 2 + num_buses + 1):
    val = sheet1.cell(row=i, column=10).value
    if val in [1,2,3]:
        demand_nodes[val].append(dem_index)
        dem_index += 1

pg_var = {i: cp.Variable(nonneg=True, name=f"Seller_{i}") for i in range(len(Cg))}
pd_var = {i: cp.Variable(nonneg=True, name=f"Buyer_{i}") for i in range(len(Bd))}
theta_vars = {i: cp.Variable(name=f"Theta_{i}") for i in range(num_buses)}

constraints = []

for i, pg_max_val in enumerate(Pg_max):
    constraints += [0 <= pg_var[i], pg_var[i] <= pg_max_val]
for i, pd_max_val in enumerate(Pd_max):
    constraints += [0 <= pd_var[i], pd_var[i] <= pd_max_val]

constraints.append(theta_vars[0] == 0)

for i in range(num_buses):
    for j in connection_nodes[i+1]:
        if j > i+1:
            constraints += [
                susceptance[i+1][j] * (theta_vars[i] - theta_vars[j-1]) <= pmax[i+1][j],
                susceptance[i+1][j] * (theta_vars[i] - theta_vars[j-1]) >= -pmax[i+1][j]
            ]

equality_constraints = []
for j in range(num_buses):
    constr = cp.sum([pg_var[i] for i in generator_nodes[j+1]]) + \
             cp.sum([susceptance[j+1][i] * (theta_vars[j] - theta_vars[i-1]) for i in connection_nodes[j+1]]) - \
             cp.sum([pd_var[i] for i in demand_nodes[j+1]])
    c = constr == 0
    constraints.append(c)
    equality_constraints.append(c)

objective = cp.Maximize(
    cp.sum([pd_var[i] * Bd[i] for i in range(len(Bd))]) -
    cp.sum([pg_var[i] * Cg[i] for i in range(len(Cg))])
)

prob = cp.Problem(objective, constraints)
prob.solve()
end_time = time.time()

output_file.parent.mkdir(exist_ok=True)
wbn = Workbook()
wsn = wbn.active

wsn['F9'] = "Social Welfare (€)"
wsn['F9'].font = Font(bold=True)
wsn['F10'] = f"{round(prob.value,2)} €"
wsn['F10'].font = Font(color="FF0000")

wsn['A1'] = "Generators (MW)"
wsn['A1'].font = Font(bold=True)
for i, val in enumerate([pg_var[i].value for i in range(len(pg_var))]):
    wsn.cell(row=i+2, column=1, value=f"{round(val,2)} MW")

wsn['B1'] = "Demands (MW)"
wsn['B1'].font = Font(bold=True)
for i, val in enumerate([pd_var[i].value for i in range(len(pd_var))]):
    wsn.cell(row=i+2, column=2, value=f"{round(val,2)} MW")

wsn['A7'] = "Price for each bus (€/MWh)"
wsn['A7'].font = Font(bold=True)
for j in range(num_buses):
    wsn.cell(row=7, column=j+2, value=f"BUS {j+1}")
    wsn.cell(row=8, column=j+2, value=f"{round(abs(equality_constraints[j].dual_value), 2)} €/MWh")

wsn['D1'] = "Powerflow on each line"
wsn['D1'].font = Font(bold=True)
wsn['D2'] = "Transmission lines"
wsn['D2'].font = Font(bold=True)
wsn['D3'] = "Powerflow (MW)"
wsn['D3'].font = Font(bold=True)

y = 5
for i in range(num_buses):
    for j in connection_nodes[i+1]:
        if j > i+1:
            flow = round(abs(susceptance[i+1][j] * (theta_vars[i].value - theta_vars[j-1].value)), 2)
            wsn.cell(row=2, column=y, value=f"{i+1}-{j}").alignment = Alignment(horizontal='center')
            wsn.cell(row=3, column=y, value=f"{flow} MW").alignment = Alignment(horizontal='center')
            y += 1

CS = 0
for i in range(num_buses):
    for j in connection_nodes[i + 1]:
        if j > i + 1:
            CS += (susceptance[i + 1][j] * (theta_vars[i].value - theta_vars[j - 1].value)) * (
                        equality_constraints[j - 1].dual_value - equality_constraints[i].dual_value)

wsn['F7'] = "Congestion Surplus (€)"
wsn['F7'].font = Font(bold=True)
wsn['F8'] = f"{abs(CS):.1f} €"
wsn['F8'].font = Font(color="FF0000")

wsn['F6'] = f"Solver Execution Time: {round(end_time-start_time,2)} sec"

wbn.save(output_file)
print(f"Results saved to {output_file}")
